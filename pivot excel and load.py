# -*- coding: utf-8 -*-
"""
Created on Wed Mar 13 12:22:42 2019

@author: AVariyan
"""
import openpyxl as xl
import pandas as pd
from openpyxl.utils import get_column_letter
import os
# =============================================================================
# from sqlalchemy import create_engine
# import urllib
# =============================================================================


class pivotXL_and_Load:
    def pivot_and_load(self):
# =============================================================================
#         quoted = urllib.parse.quote_plus("DRIVER={SQL Server Native Client 10.0};SERVER=INCHN-SQLDB01\BODSDev;DATABASE=Tableau;UID=Tableau;PWD=olam@123") 
#         engine = create_engine('mssql+pyodbc:///?odbc_connect={}'.format(quoted))
#         
# =============================================================================
        def xlref(row, column, zero_indexed=True):
            if zero_indexed:
                row += 1
                column += 1
            return get_column_letter(column) + str(row)
        sheets=['Yearly','Quarterly'] #sheets you want to process
        os.chdir("path") 
        # Get the active worksheet
        for sheet in sheets:
        
            workbook = xl.load_workbook(filename="File1.xlsx",data_only=True,read_only=True)
            # Get the worksheet
            ws = workbook[sheet]
         
            row_itr_start=0
            col_itr_start=0
            row_itr_end=ws.max_row-1
            col_itr_end=ws.max_column-1
            
            
            iter_start=xlref(row_itr_start,col_itr_start)
            íter_end=xlref(row_itr_end,col_itr_end)
            data_rows = []
            for row in ws[iter_start:íter_end]:
                data_cols = []
                for cell in row:
                    #print(cell.value)
                    data_cols.append(cell.value)
                data_rows.append(data_cols)
            
            
            df = pd.DataFrame(data_rows)
            
            headers = df.iloc[0]
            new_df  = pd.DataFrame(df.values[1:], columns=headers)
            new_df['Particulars']=new_df['Particulars'].str.strip()
            pivot_df=new_df.pivot_table(new_df,index=[column_list]	,columns='column to pivot',aggfunc='first',fill_value=0) #mention the default columns as index and column_list
            pivot_df.head()
            pivot_df=pivot_df.stack(0)
            pivot_df = pivot_df.reset_index()
            pivot_df=pivot_df.drop(0,axis=1)
            pivot_df.to_excel("Target filename",index=False)
               
if __name__ == '__main__':
    pivotXL_and_Load = pivotXL_and_Load()
    pivotXL_and_Load.pivot_and_load()
